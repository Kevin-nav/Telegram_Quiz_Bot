from __future__ import annotations

import re
import sys
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import create_engine, text

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from src.core.config import get_settings
from src.domains.catalog.data import (
    FACULTIES,
    LEVELS,
    PROGRAM_COURSES,
    SEMESTERS,
    _TIMETABLE_PREFIX_PROGRAMS,
)


DEFAULT_TIMETABLE_PATH = (
    Path(__file__).resolve().parents[2]
    / "timetable"
    / "Final_Examination_Timetable_Semester_One_2025_2026_Academic_Year.xlsx"
)
DEFAULT_OUTPUT_PATH = (
    Path(__file__).resolve().parents[1]
    / "docs"
    / "academic_catalog_from_timetable.md"
)
CLASS_PREFIX_PATTERN = re.compile(r"^([A-Z]{2})\s*\d")
PREFIX_HINTS = {
    "ET": "Unmapped in backend catalog; appears to be an Electrical/Electronic Technology class prefix from the timetable.",
    "PM": "Unmapped in backend catalog; handbook lists Diploma in Plant and Maintenance Engineering under SRID.",
    "TC": "Unmapped in backend catalog; programme/faculty not confirmed from the current handbook extraction.",
}


def load_question_counts() -> dict[str, int]:
    settings = get_settings()
    engine = create_engine(settings.sync_database_url, pool_pre_ping=True)
    with engine.connect() as conn:
        rows = conn.execute(
            text(
                """
                SELECT course_slug, COUNT(*) AS ready_rows
                FROM question_bank
                WHERE status = 'ready'
                GROUP BY course_slug
                """
            )
        ).fetchall()
    return {course_slug: int(ready_rows) for course_slug, ready_rows in rows}


def load_timetable_prefix_samples(
    timetable_path: Path,
) -> dict[str, dict[str, object]]:
    workbook = load_workbook(timetable_path, read_only=True, data_only=True)
    prefix_samples: dict[str, dict[str, object]] = {}

    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        for row in worksheet.iter_rows(min_row=4, values_only=True):
            course_no = row[1]
            course_name = row[2]
            class_name = row[3]
            if not class_name or not course_name:
                continue

            match = CLASS_PREFIX_PATTERN.match(
                " ".join(str(class_name).strip().upper().split())
            )
            if not match:
                continue

            prefix = match.group(1)
            details = prefix_samples.setdefault(
                prefix,
                {
                    "count": 0,
                    "samples": [],
                },
            )
            details["count"] = int(details["count"]) + 1
            samples = details["samples"]
            if len(samples) < 5:
                samples.append(f"{class_name} | {course_no} | {course_name}")

    return prefix_samples


def build_program_lookup() -> dict[str, str]:
    return {
        program["code"]: program["name"]
        for faculty in FACULTIES
        for program in faculty["programs"]
    }


def format_catalog_markdown(
    *,
    question_counts: dict[str, int],
    prefix_samples: dict[str, dict[str, object]],
) -> str:
    lines: list[str] = [
        "# Academic Catalog from Timetable",
        "",
        "Sources:",
        "- `timetable/Final_Examination_Timetable_Semester_One_2025_2026_Academic_Year.xlsx`",
        "- `timetable/The Students Handbook 2026_260326_134009.pdf`",
        "- `src/domains/catalog/data.py`",
        "",
        "This file lists the current backend faculty/programme/course catalog and marks each course with the number of ready question-bank rows currently available for that course slug.",
        "",
    ]

    level_names = {level["code"]: level["name"] for level in LEVELS}
    semester_names = {semester["code"]: semester["name"] for semester in SEMESTERS}

    for faculty in FACULTIES:
        lines.extend([f"## {faculty['name']} (`{faculty['code']}`)", ""])
        for program in faculty["programs"]:
            program_code = program["code"]
            lines.extend([f"### {program['name']} (`{program_code}`)", ""])

            courses_by_section: dict[tuple[str, str], list[dict]] = defaultdict(list)
            for course in PROGRAM_COURSES.get(program_code, []):
                courses_by_section[
                    (course["level_code"], course["semester_code"])
                ].append(course)

            if not courses_by_section:
                lines.extend(["No seeded courses yet.", ""])
                continue

            for level_code, semester_code in sorted(courses_by_section):
                level_name = level_names.get(level_code, f"Level {level_code}")
                semester_name = semester_names.get(
                    semester_code,
                    semester_code.title(),
                )
                lines.extend([f"#### {level_name} - {semester_name}", ""])
                for course in sorted(
                    courses_by_section[(level_code, semester_code)],
                    key=lambda item: item["name"],
                ):
                    course_code = course["code"]
                    ready_rows = question_counts.get(course_code, 0)
                    lines.append(
                        f"- {course['name']} (`{course_code}`) - ready questions: {ready_rows}"
                    )
                lines.append("")

    lines.extend(["## Timetable Prefix Mapping Audit", ""])
    program_lookup = build_program_lookup()
    for prefix in sorted(prefix_samples):
        mapped_program_code = _TIMETABLE_PREFIX_PROGRAMS.get(prefix)
        if mapped_program_code:
            mapped_program_name = program_lookup.get(
                mapped_program_code,
                mapped_program_code,
            )
            lines.append(
                f"- `{prefix}` -> {mapped_program_name} (`{mapped_program_code}`), timetable rows: {prefix_samples[prefix]['count']}"
            )
            continue

        hint = PREFIX_HINTS.get(prefix, "Unmapped in backend catalog.")
        lines.append(
            f"- `{prefix}` -> {hint} Timetable rows: {prefix_samples[prefix]['count']}; examples: {', '.join(prefix_samples[prefix]['samples'])}"
        )

    lines.append("")
    return "\n".join(lines)


def main() -> int:
    question_counts = load_question_counts()
    prefix_samples = load_timetable_prefix_samples(DEFAULT_TIMETABLE_PATH)
    markdown = format_catalog_markdown(
        question_counts=question_counts,
        prefix_samples=prefix_samples,
    )
    DEFAULT_OUTPUT_PATH.write_text(markdown, encoding="utf-8")
    print(f"Wrote {DEFAULT_OUTPUT_PATH}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
